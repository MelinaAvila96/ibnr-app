"""
exports.py
----------
Build Excel and PDF export files from IBNR results.

Functions
---------
build_excel() — multi-sheet Excel workbook
build_pdf()   — formatted PDF report using reportlab
"""

import io
import numpy as np
import pandas as pd
from datetime import date


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def build_excel(results, triangle, fdi_table, exclusions,
                premium_total=None, meta=None, latest_label="Latest Paid"):
    """
    Build a multi-sheet Excel workbook.

    Sheets
    ------
    1. Summary          — total IBNR and ultimate LR per method
    2. IBNR Comparison  — IBNR by accident period and method
    3. Triangle         — cumulative loss triangle
    4. FDI Factors      — individual link ratios + simple average
    5. Exclusion Log    — excluded cells and comments (if any)
    6. <Method name>    — full detail per method (one sheet each)

    Parameters
    ----------
    premium_total : float, optional
        Earned premium over the triangle's accident periods. When given, the
        summary shows the true Ultimate Loss Ratio (ultimate / premium);
        otherwise the column falls back to Ultimate / Latest Paid and is
        labelled accordingly.
    meta : dict, optional
        Run context ({"Base": ..., "Tail factor": ..., ...}) printed under
        the report title so the file documents how it was produced.
    latest_label : str
        What the latest diagonal is: "Latest Paid" (default) or
        "Latest Incurred" when the triangle base includes case reserves.

    Returns
    -------
    bytes
        Excel file as bytes (ready for st.download_button).
    """
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from app.methods import compute_fdi_avg, compute_cdfs, summarize_results

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # --- Styles ---
    HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
    SUBHEAD_FILL  = PatternFill("solid", fgColor="2E75B6")
    TOTAL_FILL    = PatternFill("solid", fgColor="D6E4F0")
    WARN_FILL     = PatternFill("solid", fgColor="FFF3CD")
    EXCL_FILL     = PatternFill("solid", fgColor="E2E3E5")

    WHITE_FONT    = Font(color="FFFFFF", bold=True, size=11)
    BOLD_FONT     = Font(bold=True, size=11)
    NORMAL_FONT   = Font(size=11)
    TOTAL_FONT    = Font(bold=True, size=11, color="1F4E79")
    CENTER        = Alignment(horizontal="center", vertical="center")
    RIGHT         = Alignment(horizontal="right")

    thin = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    NUM_FMT  = '#,##0'
    DEC_FMT  = '#,##0.00'
    PCT_FMT  = '0.0%'
    FDI_FMT  = '0.0000'

    def style_header(cell, wide=False):
        cell.fill      = HEADER_FILL
        cell.font      = WHITE_FONT
        cell.alignment = CENTER
        cell.border    = BORDER

    def style_subheader(cell):
        cell.fill      = SUBHEAD_FILL
        cell.font      = WHITE_FONT
        cell.alignment = CENTER
        cell.border    = BORDER

    def style_total(cell):
        cell.fill   = TOTAL_FILL
        cell.font   = TOTAL_FONT
        cell.border = BORDER

    def autofit(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 30)

    # ------------------------------------------------------------------ #
    # Sheet 1 — Summary
    # ------------------------------------------------------------------ #
    ws = wb.create_sheet("Summary")
    ws.append(["IBNR Reserve Estimation — Summary"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Generated: {date.today().strftime('%B %d, %Y')}"])
    for k, v in (meta or {}).items():
        ws.append([f"{k}: {v}"])
        ws.cell(ws.max_row, 1).font = Font(size=10, color="666666")
    ws.append([])

    # With premium, show the true loss ratio; without it, don't mislabel the
    # ultimate-to-diagonal ratio as a loss ratio.
    ratio_header = ("Ultimate Loss Ratio" if premium_total
                    else f"Ultimate / {latest_label}")
    headers = ["Method", f"Total {latest_label}", "Total IBNR", "Total Ultimate", ratio_header, "ELR"]
    ws.append(headers)
    header_row = ws.max_row
    for i, h in enumerate(headers, 1):
        style_header(ws.cell(header_row, i))

    latest_total = float(np.nansum(results[0].latest_paid))
    for r in results:
        if premium_total:
            ratio = r.total_ultimate / premium_total
        else:
            ratio = r.total_ultimate / latest_total if latest_total else 0
        row = [
            r.method,
            round(latest_total),
            round(r.total_ibnr),
            round(r.total_ultimate),
            ratio,
            round(r.elr, 4) if r.elr is not None else "—",
        ]
        ws.append(row)
        last_row = ws.max_row
        ws.cell(last_row, 2).number_format = NUM_FMT
        ws.cell(last_row, 3).number_format = NUM_FMT
        ws.cell(last_row, 4).number_format = NUM_FMT
        ws.cell(last_row, 5).number_format = PCT_FMT
        ws.cell(last_row, 6).number_format = FDI_FMT

    autofit(ws)

    # ------------------------------------------------------------------ #
    # Sheet 2 — IBNR Comparison
    # ------------------------------------------------------------------ #
    ws2 = wb.create_sheet("IBNR Comparison")
    comp = summarize_results(results)
    method_names = [r.method for r in results]

    ws2.append(["Accident Period"] + method_names)
    for i, h in enumerate(["Accident Period"] + method_names, 1):
        style_header(ws2.cell(1, i))

    for idx, row_data in comp.iterrows():
        row = [str(idx)] + [round(v, 1) if not pd.isna(v) else None for v in row_data]
        ws2.append(row)
        last_row = ws2.max_row
        is_total = str(idx) == "TOTAL"
        for col_i in range(2, len(row) + 1):
            cell = ws2.cell(last_row, col_i)
            cell.number_format = NUM_FMT
            if is_total:
                style_total(cell)

    autofit(ws2)

    # ------------------------------------------------------------------ #
    # Sheet 3 — Cumulative Triangle
    # ------------------------------------------------------------------ #
    ws3 = wb.create_sheet("Triangle")
    dev_cols = triangle.columns.tolist()

    ws3.append(["Accident Period"] + [f"{d}m" for d in dev_cols])
    for i, h in enumerate(["Accident Period"] + [f"{d}m" for d in dev_cols], 1):
        style_header(ws3.cell(1, i))

    for period in triangle.index:
        row = [period] + [
            round(triangle.loc[period, d]) if not pd.isna(triangle.loc[period, d]) else None
            for d in dev_cols
        ]
        ws3.append(row)
        last_row = ws3.max_row
        for col_i in range(2, len(row) + 1):
            ws3.cell(last_row, col_i).number_format = NUM_FMT

    autofit(ws3)

    # ------------------------------------------------------------------ #
    # Sheet 4 — FDI Factors
    # ------------------------------------------------------------------ #
    ws4 = wb.create_sheet("FDI Factors")
    from app.methods import detect_anomalies
    flags = detect_anomalies(fdi_table)
    fdi_avg = compute_fdi_avg(fdi_table, exclusions)

    ws4.append(["Accident Period"] + fdi_table.columns.tolist())
    for i, h in enumerate(["Accident Period"] + fdi_table.columns.tolist(), 1):
        style_header(ws4.cell(1, i))

    for period in fdi_table.index:
        row = [period]
        for col in fdi_table.columns:
            v = fdi_table.loc[period, col]
            row.append(round(v, 4) if not pd.isna(v) else None)
        ws4.append(row)
        last_row = ws4.max_row

        for j, col in enumerate(fdi_table.columns, 2):
            cell = ws4.cell(last_row, j)
            cell.number_format = FDI_FMT
            flag = flags.loc[period, col]
            is_excl = (period, col) in exclusions
            if is_excl:
                cell.fill = EXCL_FILL
                cell.font = Font(color="888888", size=11, strike=True)
            elif flag == "warning":
                cell.fill = WARN_FILL

    # Average row
    avg_row = ["Simple average"] + [
        round(fdi_avg[c], 4) if not pd.isna(fdi_avg[c]) else None
        for c in fdi_table.columns
    ]
    ws4.append(avg_row)
    last_row = ws4.max_row
    for i, _ in enumerate(avg_row, 1):
        style_total(ws4.cell(last_row, i))
        ws4.cell(last_row, i).number_format = FDI_FMT if i > 1 else "General"

    autofit(ws4)

    # ------------------------------------------------------------------ #
    # Sheet 5 — Exclusion Log
    # ------------------------------------------------------------------ #
    ws5 = wb.create_sheet("Exclusion Log")
    ws5.append(["Accident Period", "Development Transition", "FDI Value", "Comment"])
    for i in range(1, 5):
        style_header(ws5.cell(1, i))

    if exclusions:
        for (period, col), comment in exclusions.items():
            val = fdi_table.loc[period, col] if period in fdi_table.index and col in fdi_table.columns else None
            ws5.append([period, col, round(val, 4) if val is not None and not pd.isna(val) else None, comment])
    else:
        ws5.append(["No exclusions applied.", None, None, None])

    autofit(ws5)

    # ------------------------------------------------------------------ #
    # Sheet per method — full detail
    # ------------------------------------------------------------------ #
    for r in results:
        sheet_name = r.method[:30]
        wsm = wb.create_sheet(sheet_name)
        df_detail = r.to_dataframe()

        cols = df_detail.columns.tolist()
        latest_col_name = latest_label.lower().replace(" ", "_")
        display_cols = [latest_col_name if c == "latest_paid" else c for c in cols]
        wsm.append(["Accident Period"] + display_cols)
        for i, h in enumerate(["Accident Period"] + display_cols, 1):
            style_header(wsm.cell(1, i))

        # cdf/elr are factors and pct_reported a percentage — an integer
        # format would display e.g. a 1.0530 CDF as "1".
        col_fmts = {"cdf": FDI_FMT, "elr": FDI_FMT, "pct_reported": PCT_FMT}
        for idx, row_data in df_detail.iterrows():
            row = [str(idx)] + [
                round(v, 4) if isinstance(v, float) and not pd.isna(v)
                else (None if isinstance(v, float) and pd.isna(v) else v)
                for v in row_data
            ]
            wsm.append(row)
            last_row = wsm.max_row
            is_total = str(idx) == "TOTAL"
            for col_i, col_name in enumerate(cols, 2):
                cell = wsm.cell(last_row, col_i)
                cell.number_format = col_fmts.get(col_name, NUM_FMT)
                if is_total:
                    style_total(cell)

        autofit(wsm)

    # --- Save ---
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# PDF export
# ---------------------------------------------------------------------------

def build_pdf(results, triangle, fdi_table, exclusions,
              premium_total=None, meta=None, latest_label="Latest Paid"):
    """
    Build a formatted PDF report.

    Parameters
    ----------
    premium_total : float, optional
        Earned premium over the triangle's periods — enables the true
        Ultimate Loss Ratio in the summary (see build_excel).
    meta : dict, optional
        Run context printed on the cover (base, tail, segment, ...).

    Returns
    -------
    bytes
        PDF file as bytes.
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph,
        Spacer, HRFlowable, PageBreak,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from app.methods import compute_fdi_avg, summarize_results

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=2 * cm,    bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    NAVY  = colors.HexColor("#1F4E79")
    BLUE  = colors.HexColor("#2E75B6")
    LIGHT = colors.HexColor("#D6E4F0")
    WARN  = colors.HexColor("#FFF3CD")
    GREY  = colors.HexColor("#E2E3E5")

    title_style = ParagraphStyle("title", fontSize=18, textColor=NAVY,
                                  fontName="Helvetica-Bold", spaceAfter=4)
    h2_style    = ParagraphStyle("h2",    fontSize=13, textColor=NAVY,
                                  fontName="Helvetica-Bold", spaceBefore=14, spaceAfter=6)
    body_style  = styles["Normal"]
    caption_style = ParagraphStyle("cap", fontSize=9, textColor=colors.grey,
                                    spaceAfter=6)

    def table_style_base(header_rows=1):
        return TableStyle([
            ("BACKGROUND",  (0, 0), (-1, header_rows - 1), NAVY),
            ("TEXTCOLOR",   (0, 0), (-1, header_rows - 1), colors.white),
            ("FONTNAME",    (0, 0), (-1, header_rows - 1), "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, -1), 9),
            ("ALIGN",       (1, 0), (-1, -1), "RIGHT"),
            ("ALIGN",       (0, 0), (0, -1), "LEFT"),
            ("ROWBACKGROUNDS", (0, header_rows), (-1, -1),
             [colors.white, colors.HexColor("#F0F6FC")]),
            ("GRID",        (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
            ("TOPPADDING",  (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ])

    def fmt(v, decimals=0):
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return "—"
        if decimals == 0:
            return f"{v:,.0f}"
        return f"{v:,.{decimals}f}"

    story = []

    # --- Cover ---
    story.append(Paragraph("IBNR Reserve Estimation Report", title_style))
    story.append(Paragraph(f"Generated: {date.today().strftime('%B %d, %Y')}", caption_style))
    if meta:
        story.append(Paragraph(
            " · ".join(f"{k}: {v}" for k, v in meta.items()), caption_style,
        ))
    story.append(HRFlowable(width="100%", thickness=1, color=NAVY, spaceAfter=12))

    # --- Summary table ---
    story.append(Paragraph("Summary", h2_style))

    latest_total = float(np.nansum(results[0].latest_paid))
    ratio_header = "Loss Ratio" if premium_total else f"Ultimate / {latest_label}"
    sum_data = [["Method", latest_label, "Total IBNR", "Ultimate", ratio_header, "ELR"]]
    for r in results:
        if premium_total:
            ratio = r.total_ultimate / premium_total
        else:
            ratio = r.total_ultimate / latest_total if latest_total else 0
        sum_data.append([
            r.method,
            fmt(latest_total),
            fmt(r.total_ibnr),
            fmt(r.total_ultimate),
            f"{ratio:.1%}",
            f"{r.elr:.4f}" if r.elr is not None else "—",
        ])

    t = Table(sum_data, repeatRows=1)
    t.setStyle(table_style_base())
    story.append(t)
    story.append(Spacer(1, 12))

    # --- IBNR comparison ---
    story.append(Paragraph("IBNR by Accident Period", h2_style))
    comp = summarize_results(results)
    method_names = [r.method for r in results]

    comp_data = [["Accident Period"] + method_names]
    for idx, row in comp.iterrows():
        comp_data.append([str(idx)] + [fmt(v) for v in row])

    t2 = Table(comp_data, repeatRows=1)
    st2 = table_style_base()
    # Bold total row
    total_row_idx = len(comp_data) - 1
    st2.add("BACKGROUND", (0, total_row_idx), (-1, total_row_idx), LIGHT)
    st2.add("FONTNAME",   (0, total_row_idx), (-1, total_row_idx), "Helvetica-Bold")
    t2.setStyle(st2)
    story.append(t2)
    story.append(Spacer(1, 12))

    # --- FDI table ---
    story.append(PageBreak())
    story.append(Paragraph("Individual FDI Factors", h2_style))
    story.append(Paragraph(
        "Highlighted cells: yellow = anomaly (moderate IQR outlier), "
        "grey strikethrough = excluded by user.",
        caption_style
    ))

    from app.methods import detect_anomalies
    flags = detect_anomalies(fdi_table)
    fdi_avg = compute_fdi_avg(fdi_table, exclusions)

    fdi_data = [["AY"] + fdi_table.columns.tolist()]
    fdi_cell_styles = []

    for row_i, period in enumerate(fdi_table.index, 1):
        row = [str(period)]
        for col_j, col in enumerate(fdi_table.columns, 1):
            v    = fdi_table.loc[period, col]
            flag = flags.loc[period, col]
            is_excl = (period, col) in exclusions

            row.append(f"{v:.4f}" if not pd.isna(v) else "—")

            if is_excl:
                fdi_cell_styles.append(("BACKGROUND", (col_j, row_i), (col_j, row_i), GREY))
                fdi_cell_styles.append(("TEXTCOLOR",  (col_j, row_i), (col_j, row_i), colors.grey))
            elif flag == "warning":
                fdi_cell_styles.append(("BACKGROUND", (col_j, row_i), (col_j, row_i), WARN))

        fdi_data.append(row)

    # Average row
    avg_row = ["Simple avg"] + [
        f"{fdi_avg[c]:.4f}" if not pd.isna(fdi_avg[c]) else "—"
        for c in fdi_table.columns
    ]
    fdi_data.append(avg_row)
    avg_row_idx = len(fdi_data) - 1
    fdi_cell_styles += [
        ("BACKGROUND", (0, avg_row_idx), (-1, avg_row_idx), LIGHT),
        ("FONTNAME",   (0, avg_row_idx), (-1, avg_row_idx), "Helvetica-Bold"),
    ]

    t3 = Table(fdi_data, repeatRows=1)
    st3 = table_style_base()
    for s in fdi_cell_styles:
        st3.add(*s)
    t3.setStyle(st3)
    story.append(t3)

    # --- Exclusion log ---
    if exclusions:
        story.append(Spacer(1, 16))
        story.append(Paragraph("Exclusion Log", h2_style))
        excl_data = [["Accident Period", "Transition", "FDI Value", "Reason"]]
        for (period, col), comment in exclusions.items():
            v = fdi_table.loc[period, col] if period in fdi_table.index else None
            excl_data.append([
                str(period), col,
                f"{v:.4f}" if v is not None and not np.isnan(v) else "—",
                comment,
            ])
        t4 = Table(excl_data, repeatRows=1, colWidths=[3 * cm, 3 * cm, 3 * cm, None])
        t4.setStyle(table_style_base())
        story.append(t4)

    # --- Per-method detail ---
    for r in results:
        story.append(PageBreak())
        story.append(Paragraph(f"Method: {r.method}", h2_style))
        if r.elr is not None:
            story.append(Paragraph(f"Expected Loss Ratio (ELR): {r.elr:.4f} ({r.elr:.1%})",
                                    caption_style))

        df_d = r.to_dataframe()
        cols = df_d.columns.tolist()

        def fmt_detail(col_name, v):
            if not isinstance(v, float):
                return str(v)
            if v is None or np.isnan(v):
                return "—"
            if col_name in ("cdf", "elr"):
                return f"{v:.4f}"
            if col_name == "pct_reported":
                return f"{v:.1%}"
            return fmt(v, 2)

        latest_col_name = latest_label.lower().replace(" ", "_")
        display_cols = [latest_col_name if c == "latest_paid" else c for c in cols]
        detail_data = [["Accident Period"] + display_cols]
        for idx, row in df_d.iterrows():
            detail_data.append(
                [str(idx)] + [fmt_detail(c, v) for c, v in zip(cols, row)]
            )

        t5 = Table(detail_data, repeatRows=1)
        st5 = table_style_base()
        total_i = len(detail_data) - 1
        st5.add("BACKGROUND", (0, total_i), (-1, total_i), LIGHT)
        st5.add("FONTNAME",   (0, total_i), (-1, total_i), "Helvetica-Bold")
        t5.setStyle(st5)
        story.append(t5)

    doc.build(story)
    buf.seek(0)
    return buf.read()
