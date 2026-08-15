"""
test_exports.py
---------------
Checks on the Excel export: loss-ratio labelling/denominator and factor
number formats (findings B1/B2).

Run with:
    python -m unittest tests/test_exports.py -v
"""

import io
import unittest
import numpy as np
import pandas as pd
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.methods import chain_ladder, compute_fdi
from app.exports import build_excel


def _triangle():
    data = np.array([
        [1000, 1800, 2300, 2500],
        [1100, 1950, 2450, np.nan],
        [1200, 2100, np.nan, np.nan],
        [1300, np.nan, np.nan, np.nan],
    ], dtype=float)
    return pd.DataFrame(data, index=[2020, 2021, 2022, 2023],
                        columns=[12, 24, 36, 48])


class TestExcelExport(unittest.TestCase):

    def setUp(self):
        self.tri = _triangle()
        self.fdi = compute_fdi(self.tri)
        self.res = chain_ladder(self.tri)

    def _load(self, **kwargs):
        from openpyxl import load_workbook
        raw = build_excel([self.res], self.tri, self.fdi, {}, **kwargs)
        return load_workbook(io.BytesIO(raw))

    def _summary_table(self, ws):
        """Return (header_row_values, first_data_row_values) of the summary."""
        rows = list(ws.iter_rows(values_only=True))
        for i, row in enumerate(rows):
            if row and row[0] == "Method":
                return list(row), list(rows[i + 1])
        raise AssertionError("Summary header row not found")

    def test_loss_ratio_uses_premium_when_given(self):
        premium_total = 20000.0
        wb = self._load(premium_total=premium_total,
                        meta={"Base": "Paid", "Tail factor": "1.000"})
        header, data = self._summary_table(wb["Summary"])
        ratio_col = header.index("Ultimate Loss Ratio")
        self.assertAlmostEqual(data[ratio_col],
                               self.res.total_ultimate / premium_total,
                               places=6)

    def test_ratio_relabelled_without_premium(self):
        wb = self._load()
        header, data = self._summary_table(wb["Summary"])
        self.assertNotIn("Ultimate Loss Ratio", header)
        ratio_col = header.index("Ultimate / Latest Paid")
        latest_total = float(np.nansum(self.res.latest_paid))
        self.assertAlmostEqual(data[ratio_col],
                               self.res.total_ultimate / latest_total,
                               places=6)

    def test_meta_written_on_summary(self):
        wb = self._load(meta={"Base": "Incurred (paid + RSP)", "Segment": "All"})
        texts = [str(r[0]) for r in wb["Summary"].iter_rows(values_only=True) if r and r[0]]
        self.assertTrue(any("Incurred (paid + RSP)" in t for t in texts))

    def test_latest_label_for_incurred_base(self):
        wb = self._load(latest_label="Latest Incurred")
        header, _data = self._summary_table(wb["Summary"])
        self.assertIn("Total Latest Incurred", header)
        self.assertNotIn("Total Latest Paid", header)
        # Per-method detail sheet renames the diagonal column too.
        detail_header = [c.value for c in wb["Chain Ladder"][1]]
        self.assertIn("latest_incurred", detail_header)
        self.assertNotIn("latest_paid", detail_header)

    def test_cdf_column_not_integer_formatted(self):
        wb = self._load()
        ws = wb["Chain Ladder"]
        header = [c.value for c in ws[1]]
        cdf_idx = header.index("cdf") + 1          # 1-based
        cdf_cell = ws.cell(2, cdf_idx)
        self.assertEqual(cdf_cell.number_format, "0.0000")


if __name__ == "__main__":
    unittest.main(verbosity=2)
